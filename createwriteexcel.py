import random

import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active

define_range = int(input("Please enter the number of data sets you want to create\n"))

#function to just enter the column values to the excel
def enter_column_names():
    c1 = sheet.cell(row = 1, column = 1)
    c1.value = "Sl.no"
    c2 = sheet.cell(row = 1, column = 2)
    c2.value = "Principal amount"
    c3 =  sheet.cell(row = 1, column = 3)
    c3.value = "Percentage"
    c4 = sheet.cell(row = 1, column = 4)
    c4.value = "Period"
    c5 = sheet.cell(row = 1, column = 5)
    c5.value = "Total Simple Interest"
    return "Column names are entered"

#function to increment the sl.no column
def data_calculation():
    count = 0
    for count in range(1,define_range):
        g1 = sheet.cell(row = count + 1, column = 1)
        g1.value = count
        count = count + 1
    return "Serial numbers are created"

#function to perform the simple interest calculation data sets
def data_calculation_final():
    count = 0
    for count in range(1, define_range):
        v1 =  sheet.cell(row = count+1 , column = 2)
        v1.value = random.randrange(15000,50000,1500)
        v2 = sheet.cell(row = count+1 , column = 3)
        v2.value = random.randrange(5,13,1)
        v3 = sheet.cell(row = count+1 , column = 4)
        v3.value = random.randrange(5,10,1)
        v4 = sheet.cell(row = count+1 , column = 5)
        v4.value = v1.value * v2.value * v3.value * 0.01
        count = count + 1
    return "Data is entered"

#function to assign variables to all the defined functions
def call_data_function():
    final_variable = data_calculation()
    final_variable1 = enter_column_names()
    final_variable2 = data_calculation_final()
    wb.save("//Users//krishnabharadwaj//Desktop//Data_excel.xlsx")
    print("Excel created successfully")

#calling function to generate the excel
call_data_function()








